from typing import List
from pathlib import Path
from openpyxl import Workbook, load_workbook

from flightkit.utils import logger
from flightkit.core.models import Flight
from flightkit.core.interfaces import IFlightExporter

ARTIFACTS_DIR = Path(__file__).parent.parent.parent.parent / "artifacts"
ARTIFACTS_DIR.mkdir(parents=True, exist_ok=True)


class ExcelExporter(IFlightExporter):
    """
    This exporter always saves files inside artifacts folder.
    """

    HEADERS = [
        "Airline",
        "Flight Number",
        "Departure Time",
        "Arrival Time",
        "Price",
        "Capacity",
        "Origin",
        "Destination",
    ]

    def export(self, flights: List[Flight], filename: str) -> None:
        """
        Save flights into Excel under artifacts folder.
        """
        if not flights:
            return

        output_path = self._resolve_output_path(filename)
        workbook, sheet = self._get_workbook_and_sheet(output_path)

        for flight in flights:
            row_data = [
                flight.airline_name,
                flight.flight_number,
                flight.departure_time,
                flight.arrival_time,
                flight.price,
                flight.capacity,
                flight.origin,
                flight.destination,
            ]
            sheet.append(row_data)

        try:
            workbook.save(output_path)
        except PermissionError as exc:
            logger.error(f"Cannot write file. Close '{output_path.name}' and retry.")
            raise exc

    def _get_workbook_and_sheet(self, file_path: Path):
        """
        Open existing workbook or create a new one with headers.
        """
        if file_path.exists():
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(self.HEADERS)
        return workbook, sheet

    def _resolve_output_path(self, filename: str) -> Path:
        """
        Map any filename to artifacts directory.
        """
        clean_name = Path(filename or "flight_database.xlsx").name
        return ARTIFACTS_DIR / clean_name